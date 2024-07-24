import os
import re
import pandas as pd
import tempfile
from flask import Flask, request, render_template, send_file, redirect
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
import time

app = Flask(__name__)

def read_file_with_multiple_encodings(file_path, encodings=['utf-8', 'iso-8859-1', 'windows-1252']):
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                return f.read()
        except UnicodeDecodeError:
            print(f"Failed to decode with {encoding}, trying next encoding.")
    raise UnicodeDecodeError(f"All tried encodings failed for {file_path}")

def process_html_files(file_addresses, wb, summary_data):
    ws_event_log = wb.create_sheet(title="Output_event_log")

    ws_event_log.append(['IP Device', 'System Name', 'Entry ID', 'Date', 'Time', 'Error Type', 'Error Code', 'TaskName', 'Filename', 'Line', 'Parameter'])

    for index, file_address in enumerate(file_addresses, start=1):
        try:
            file_base_name = os.path.splitext(os.path.basename(file_address))[0]
            html_content = read_file_with_multiple_encodings(file_address)

            ip_device_match = re.search(r'IP=(\d+\.\d+\.\d+\.\d+)', html_content)
            ip_device = ip_device_match.group(1) if ip_device_match else "Unknown"

            system_name_match = re.search(r'System Name:\s*([^\n\r]*)', html_content)
            system_name = system_name_match.group(1).strip().split("<")[0] if system_name_match else "Unknown"

            pattern1 = re.compile(r'<tr><td>(\d+): <font color="#(?:3366FF|606060|009900)">(\d{2}\.\d{2}\.\d{2})\s*(\d{2}:\d{2}:\d{2}):\s*(\S+)\s*(\S+)\s*(\S+)\s*,\s*(\S+)\s*,\s*(\d+)<br>\s*\.+(\S+)')
            events = re.findall(pattern1, html_content)

            if not events:
                print(f"No events found in {file_address}.")
                continue

            df = pd.DataFrame(events, columns=['Entry ID', 'Date', 'Time', 'Error Type', 'Error Code', 'TaskName', 'Filename', 'Line', 'Parameter'])
            df.insert(0, 'System Name', system_name)
            df.insert(0, 'IP Device', ip_device)
            df_sorted = df.sort_values(by='Entry ID')

            for r in dataframe_to_rows(df_sorted, index=False, header=False):
                ws_event_log.append(r)
            summary_data.append([index, file_base_name, ip_device, system_name])
        except FileNotFoundError:
            print(f"File not found: {file_address}")
        except Exception as e:
            print(f"Error processing {file_address}: {e}")
        finally:
            os.remove(file_address)

    ws_event_log.auto_filter.ref = f"A1:{chr(64 + ws_event_log.max_column)}1"

    for col in ws_event_log.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        ws_event_log.column_dimensions[column].width = max_length + 2

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for cell in ws_event_log[1]:
        cell.fill = yellow_fill

def process_log_files(file_addresses, wb, summary_data):
    ws_syslog = wb.create_sheet(title="Syslog")

    ws_syslog.append(['System Name', 'Month', 'Date', 'Timestamp', 'Facility', 'Severity Level', 'Mnemonic', 'Message Text', 'Traceback'])

    for index, file_address in enumerate(file_addresses, start=1):
        try:
            system_name = os.path.basename(file_address)
            system_name = system_name.split('.')[0]
            system_name = system_name.split('_')[-1]
            txt_content = read_file_with_multiple_encodings(file_address)

            pattern1 = re.compile(r'\**(\w{2,3})\s+(\d{1,2}) (\d{2}:\d{2}:\d{2}\.\d{3}\s*\S*): %(\S+)-(\d)-(\w+): (.+)\s*(?:\s*-Traceback=(.+))?')

            events = re.findall(pattern1, txt_content)

            if not events:
                print(f"No events found in {file_address}.")
                continue

            df = pd.DataFrame(events, columns=['Month', 'Date', 'Timestamp', 'Facility', 'Severity Level', 'Mnemonic', 'Message Text', 'Traceback'])
            df.insert(0, 'System Name', system_name)
            df_sorted = df.sort_values(by='Timestamp')

            for r in dataframe_to_rows(df_sorted, index=False, header=False):
                ws_syslog.append(r)

            summary_data.append([index, os.path.basename(file_address), None, system_name])
        except FileNotFoundError:
            print(f"File not found: {file_address}")
        except Exception as e:
            print(f"Error processing {file_address}: {e}")
        finally:
            os.remove(file_address)

    ws_syslog.auto_filter.ref = f"A1:{chr(64 + ws_syslog.max_column)}1"

    for col in ws_syslog.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        ws_syslog.column_dimensions[column].width = max_length + 2

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for cell in ws_syslog[1]:
        cell.fill = yellow_fill

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        file_name = request.form['file_name']
        input_files = request.files.getlist('input_folder')

        if not input_files:
            return "No files uploaded"

        html_files = [file for file in input_files if file.filename.endswith('.html')]
        log_files = [file for file in input_files if file.filename.endswith('.log')]

        temp_dir = tempfile.gettempdir()
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Number of files"
        ws_summary.append(['Total Number of Files'])
        
        summary_data = []

        if html_files:
            ws_summary.append(['No.', 'Filename', 'IP Devices', 'System Names'])
            html_file_addresses = []
            for html_file in html_files:
                file_path = os.path.join(temp_dir, secure_filename(html_file.filename))
                html_file.save(file_path)
                html_file_addresses.append(file_path)
            process_html_files(html_file_addresses, wb, summary_data)

        if log_files:
            ws_summary.append(['No.', 'Filename', 'System Names'])
            log_file_addresses = []
            for log_file in log_files:
                file_path = os.path.join(temp_dir, secure_filename(log_file.filename))
                log_file.save(file_path)
                log_file_addresses.append(file_path)
            process_log_files(log_file_addresses, wb, summary_data)

        # Update summary sheet with collected data
        total_files = len(summary_data)
        ws_summary.cell(row=1, column=2, value=total_files)
        for row in summary_data:
            ws_summary.append(row)

        for col in ws_summary.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            ws_summary.column_dimensions[column].width = max_length + 2

        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        light_yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

        ws_summary['A1'].fill = yellow_fill
        ws_summary['A2'].fill = light_yellow_fill
        for cell in ws_summary[2]:
            cell.fill = yellow_fill

        output_file_path = os.path.join(temp_dir, f'{secure_filename(file_name)}.xlsx')

        if 'Sheet' in wb.sheetnames:
            ws = wb['Sheet']
            wb.remove(ws)
            
        wb.save(output_file_path)

        response = send_file(output_file_path, as_attachment=True, download_name=f'{secure_filename(file_name)}.xlsx')
 
        # Clean up
        # os.remove(output_file_path)
        return response

    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)
